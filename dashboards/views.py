import os
import openpyxl
import uuid
import msal
import requests
import re
import time
import hashlib
import json
from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework import status
from django.conf import settings
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.db.models import Sum, OuterRef, Subquery, FloatField
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
from payments.models import Received, Modality, ServicosPagos
from clients.models import Clients
from datetime import datetime
from .serializers import ModalityResultSerializer, SharedLinkSerializer
from .models import SharedLink, LogExport
from dotenv import load_dotenv
from io import BytesIO
import logging

load_dotenv()

# Configurar o logger
logger = logging.getLogger(__name__)

class SharedLinkCreateView(generics.CreateAPIView):
    queryset = SharedLink.objects.all()
    serializer_class = SharedLinkSerializer
    permission_classes = [IsAuthenticated]
           
    def get_access_token(self):
        client_id = os.getenv('CLIENT_ID')
        tenant_id = os.getenv('TENANT_ID')
        client_secret = os.getenv('CLIENT_SECRET')
        scope = ['https://graph.microsoft.com/.default']

        # Obter o token de acesso
        app = msal.ConfidentialClientApplication(client_id, authority=f'https://login.microsoftonline.com/{tenant_id}', client_credential=client_secret)
        result = app.acquire_token_for_client(scopes=scope) 
        
        return result 
    
    def download_excel(self, request, access_token):
        drive_id = os.getenv('DRIVE_ID')
        item_id = os.getenv('FILE_ID')
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
        
        headers = {
            "Authorization": f"Bearer {access_token}"
        }

        # Fazer a requisição GET para baixar o Excel
        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            # Gerar um nome de arquivo único com timestamp e hash
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            hash_str = hashlib.md5(response.content).hexdigest()[:8]
            file_name = f"dashboard_{timestamp}_{hash_str}.xlsm"

            # Caminho completo para salvar o arquivo
            file_path = os.path.join(settings.MEDIA_ROOT, 'exported_dashboards', file_name)

            # Certificar que o diretório existe
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            # Salvar o conteúdo do arquivo localmente
            with open(file_path, 'wb') as file:
                file.write(response.content)

            # Gerar URL para o arquivo
            file_url = f"{settings.MEDIA_URL}exported_dashboards/{file_name}"
            file_url = request.build_absolute_uri(file_url)
            # Retornar a URL para download
            return JsonResponse({"download_url": file_url}, status=200)
        
        else:
            return JsonResponse({"error": f"Erro ao gerar o dashboard: {response.json()}"}, status=response.status_code)
        
    def write_values(self, access_token, sheet_name, range_address, values, log):
        print('Atualizando planilha:', sheet_name)
        drive_id = os.getenv('DRIVE_ID')
        item_id = os.getenv('FILE_ID')
        

        # URL do endpoint da API para escrever valores em uma faixa específica
        url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/workbook/worksheets/{sheet_name}/range(address='{range_address}')"

        # Dados a serem enviados para a API
        data = {
            "values": values
        }

        # Headers de autenticação e tipo de conteúdo
        headers = {
            "Authorization": f"Bearer {access_token}",
            "Content-Type": "application/json"
        }

        # Fazer a requisição PATCH para atualizar os dados no intervalo especificado
        response = requests.patch(url, headers=headers, json=data)

        if response.status_code == 200:         
            print('Dados atualizados com sucesso! Planilha: ', sheet_name)
            current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log += f'[{current_time}] - Dados atualizados com sucesso! Planilha: {sheet_name} \n'
            time.sleep(5)
            return 0, log
        else:
            print(f'Erro ao atualizar os dados: {response.json()}')
            current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log += f'[{current_time}] - Erro ao atualizar os dados: {response.json()} \n'
            return 1, log

    def perform_create(self, serializer):
        # Gera um código único para o code_url
        unique_code = str(uuid.uuid4())  # Gera um UUID como código único
        serializer.save(created_by=self.request.user, code_url=unique_code)

    def create(self, request, *args, **kwargs):
        try:
            # Chama o método de criação padrão
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            code_url = serializer.instance.code_url
            pot_econ_acao_01 = serializer.instance.info_action_01
            pot_econ_acao_02 = serializer.instance.info_action_02
            pot_econ_acao_03 = serializer.instance.info_action_03
            pot_econ_01 = serializer.instance.action_01
            pot_econ_02 = serializer.instance.action_02
            pot_econ_03 = serializer.instance.action_03
                    
            # Obter o objeto criado para usar nas operações a seguir
            user = request.user 
            shared = serializer.instance
            client_id = shared.client.pk
            fantasy_name = shared.client.fantasy_name
            info_adicional = shared.info
            date_start = shared.date_start
            date_end = shared.date_end
            
            # Montar o log
            current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_content = f'[{current_time}] - Iniciando a exportação do cliente: {fantasy_name} \n'
            
            # Formatar as datas no padrão desejado
            date_start_formatted = date_start.strftime("%d/%m/%Y")
            date_end_formatted = date_end.strftime("%d/%m/%Y")
            periodo_apuracao = f"Período de Apuração \n{date_start_formatted} - {date_end_formatted}"

            # Filtrar registros de acordo com o client_id e intervalo de data
            queryset = Received.objects.filter(
                client_id=client_id,
                data_pagamento__range=(date_start, date_end)
            ).values('modality').annotate(
                valor_bruto_sum=Sum('valor_bruto'),
                valor_taxa_sum=Sum('valor_taxa')
            )          

            # Obter informações adicionais
            client = get_object_or_404(Clients, id=client_id)
            fantasy_name = client.fantasy_name        

            received_data_view = ReceivedDataView()
            modality_data = received_data_view.get_modality_numbers(client_id, date_start, date_end)
            info_numbers = received_data_view.get_info_numbers(queryset, client_id, date_start, date_end)
            tipo_cartoes_final = received_data_view.get_tipo_cartoes(client_id, date_start, date_end)
            adquirente_data = received_data_view.get_adquirente(client_id, date_start, date_end)
            servicos_adicionais_pagos = received_data_view.get_servicos_adicionais_pagos(client_id, date_start, date_end)          
            quantidade_total_vendas = received_data_view.get_total_vendas(client_id, date_start, date_end)          

            # Calcular o valor bruto total para o client_id e intervalo de datas
            total_venda_bruta = Received.objects.filter(
                client_id=client_id,
                data_pagamento__range=(date_start, date_end)
            ).aggregate(total=Sum('valor_bruto'))['total'] or 0

            # Calcular venda bruta por tipo de cartão
            venda_por_tipo_cartao_queryset = (
                Received.objects.filter(
                    client_id=client_id,
                    data_pagamento__range=(date_start, date_end),
                    product__type_card__id__isnull=False
                )
                .values('product__type_card__id', 'product__type_card__name')
                .annotate(venda_bruta=Sum('valor_bruto'))
            )

            # Constrói a lista de resultados para vendaPorTipoCartao
            venda_por_tipo_cartao = [
                {
                    "id": entry['product__type_card__id'],
                    "title": entry['product__type_card__name'],
                    "value": round((entry['venda_bruta'] / total_venda_bruta) * 100, 2) if total_venda_bruta > 0 else 0
                }
                for entry in venda_por_tipo_cartao_queryset
            ]
            
            # CONECTAR AO ONE DRIVE PARA ALTERAR O EXCEL
            result = self.get_access_token()
            
            if "access_token" in result:
                log_entry = LogExport()           
                access_token = result['access_token']
                
                ##############
                # Tabela Info
                # Extrair o valor de antecipacaoDespesa e remover caracteres extras
                valor_str = info_numbers["antecipacaoDespesa"]["value"]
                valor_numero = float(re.sub(r'[^\d,]', '', valor_str).replace(',', '.'))

                # Multiplicar por 12
                impacto_anual_antecipacao = valor_numero * 12            
                sheet_name = "Info"
                range_address = "B1:B4"  # Ajuste o intervalo conforme necessário
                values = [
                    [fantasy_name],
                    [periodo_apuracao],
                    [impacto_anual_antecipacao],
                    [quantidade_total_vendas]
                ] 
                result_write_values, log_content = self.write_values(access_token, sheet_name, range_address, values, log_content)
                
                if result_write_values > 0:
                    log_entry.save_log(
                        user=user,  # Uma instância do usuário que fez a ação
                        log=log_content,
                        resultado=False,
                        client_id=client_id,  # ID do cliente relacionado
                        date_start=date_start,
                        date_end=date_end
                    ) 
                    return Response({'message': 'Ocorreu um erro ao exportar o dashboard. Tente novamente.'}, status=status.HTTP_400_BAD_REQUEST)
                                
                ##############
                # Tabela PotencialEconomia 
                sheet_name = "PotencialEconomia"  
                range_address = "A2:B4" 
                values = [
                    [pot_econ_acao_01, pot_econ_01],
                    [pot_econ_acao_02, pot_econ_02],
                    [pot_econ_acao_03, pot_econ_03]
                ]
                result_write_values, log_content = self.write_values(access_token, sheet_name, range_address, values, log_content)
                if result_write_values > 0:
                    log_entry.save_log(
                        user=user,  # Uma instância do usuário que fez a ação
                        log=log_content,
                        resultado=False,
                        client_id=client_id,  # ID do cliente relacionado
                        date_start=date_start,
                        date_end=date_end
                    )                 
                    return Response({'message': 'Ocorreu um erro ao exportar o dashboard. Tente novamente.'}, status=status.HTTP_400_BAD_REQUEST)
                
                ##############
                # Tabela TaxaEfetiva 
                sheet_name = "TaxaEfetiva"  
                range_address = "A2:B10" 
                # Montar o array values
                values = []

                # Adiciona os pares [title, value] ao array values
                for item in modality_data.data:
                    values.append([item["title"], float(item["value"])/100])  # Convertendo Decimal para float

                # Preenche com valores padrão se houver menos de 9 linhas
                while len(values) < 9:
                    values.append(["", 0.0])  # Preencher com strings vazias e zero (float)

                result_write_values, log_content = self.write_values(access_token, sheet_name, range_address, values, log_content)            
                if result_write_values > 0:
                    log_entry.save_log(
                        user=user,  # Uma instância do usuário que fez a ação
                        log=log_content,
                        resultado=False,
                        client_id=client_id,  # ID do cliente relacionado
                        date_start=date_start,
                        date_end=date_end
                    )                 
                    return Response({'message': 'Ocorreu um erro ao exportar o dashboard. Tente novamente.'}, status=status.HTTP_400_BAD_REQUEST)            
                ##############
                # Tabela Numeros
                # Extrair o valor de antecipacaoDespesa e remover caracteres extras
                valor_str = info_numbers["antecipacaoDespesa"]["value"]          
                sheet_name = "Numeros"
                range_address = "B1:B9"  # Ajuste o intervalo conforme necessário
                values = [
                    [info_numbers["semAntecipacaoVenda"]["value"]],
                    [info_numbers["semAntecipacaoDespesa"]["value"]],
                    [info_numbers["semAntecipacaoTaxa"]["value"]],
                    [info_numbers["antecipacaoVenda"]["value"]],
                    [info_numbers["antecipacaoDespesa"]["value"]],
                    [info_numbers["antecipacaoTaxa"]["value"]],
                    [info_numbers["vendaTotal"]["value"]],
                    [info_numbers["despesaTotal"]["value"]],
                    [info_numbers["taxaTotal"]["value"]]
                ] 
                result_write_values, log_content = self.write_values(access_token, sheet_name, range_address, values, log_content)
                if result_write_values > 0:
                    log_entry.save_log(
                        user=user,  # Uma instância do usuário que fez a ação
                        log=log_content,
                        resultado=False,
                        client_id=client_id,  # ID do cliente relacionado
                        date_start=date_start,
                        date_end=date_end
                    )                 
                    return Response({'message': 'Ocorreu um erro ao exportar o dashboard. Tente novamente.'}, status=status.HTTP_400_BAD_REQUEST)                        
                ##############
                # Tabela VendaDebito 
                sheet_name = "VendaDebito"  
                range_address = "A2:C30" 
                # Montar o array values
                values = []

                # Acessar os dados de "Debito"
                debito_items = tipo_cartoes_final[0]["Debito"]

                # Montar o array values com os itens de "Debito"
                for item in debito_items:
                    values.append([
                        item["name"],
                        float(item["Venda Bruta"]),
                        float(item["Taxa%"])/100
                    ])

                # Preenche com valores padrão se houver menos de 30 linhas
                while len(values) < 29:
                    values.append(["", "", ""])  
                result_write_values, log_content = self.write_values(access_token, sheet_name, range_address, values, log_content)
                if result_write_values > 0:
                    log_entry.save_log(
                        user=user,  # Uma instância do usuário que fez a ação
                        log=log_content,
                        resultado=False,
                        client_id=client_id,  # ID do cliente relacionado
                        date_start=date_start,
                        date_end=date_end
                    )                 
                    return Response({'message': 'Ocorreu um erro ao exportar o dashboard. Tente novamente.'}, status=status.HTTP_400_BAD_REQUEST)                                    
                ##############
                # Tabela VendaCredito 
                sheet_name = "VendaCredito"  
                range_address = "A2:C30" 
                # Montar o array values
                values = []

                # Acessar os dados de "Crebito"
                credito_items = tipo_cartoes_final[0]["Credito"]

                # Montar o array values com os itens de "Debito"
                for item in credito_items:
                    values.append([
                        item["name"],
                        float(item["Venda Bruta"]),
                        float(item["Taxa%"])/100
                    ])

                # Preenche com valores padrão se houver menos de 30 linhas
                while len(values) < 29:
                    values.append(["", "", ""])  
                result_write_values, log_content = self.write_values(access_token, sheet_name, range_address, values, log_content)
                if result_write_values > 0:
                    log_entry.save_log(
                        user=user,  # Uma instância do usuário que fez a ação
                        log=log_content,
                        resultado=False,
                        client_id=client_id,  # ID do cliente relacionado
                        date_start=date_start,
                        date_end=date_end
                    )                 
                    return Response({'message': 'Ocorreu um erro ao exportar o dashboard. Tente novamente.'}, status=status.HTTP_400_BAD_REQUEST)                                           
                ##############
                # Tabela VendaVoucher 
                sheet_name = "VendaVoucher"  
                range_address = "A2:C30" 
                # Montar o array values
                values = []

                # Acessar os dados de "Voucher"
                voucher_items = tipo_cartoes_final[0]["Voucher"]

                # Montar o array values com os itens de "Voucher"
                for item in voucher_items:
                    values.append([
                        item["name"],
                        float(item["Venda Bruta"]),
                        float(item["Taxa%"])/100
                    ])

                # Preenche com valores padrão se houver menos de 30 linhas
                while len(values) < 29:
                    values.append(["", "", ""])  
                result_write_values, log_content = self.write_values(access_token, sheet_name, range_address, values, log_content)
                if result_write_values > 0:
                    log_entry.save_log(
                        user=user,  # Uma instância do usuário que fez a ação
                        log=log_content,
                        resultado=False,
                        client_id=client_id,  # ID do cliente relacionado
                        date_start=date_start,
                        date_end=date_end
                    )                 
                    return Response({'message': 'Ocorreu um erro ao exportar o dashboard. Tente novamente.'}, status=status.HTTP_400_BAD_REQUEST)                                                
                ##############
                # Tabela VendaAdquirente 
                sheet_name = "VendaAdquirente"  
                range_address = "A2:C25" 
                # Montar o array values
                values = []

                for item in adquirente_data:
                    values.append([
                        item["name"],
                        float(item["Venda Bruta"]),
                        float(item["Taxa%"])/100
                    ])

                # Preenche com valores padrão se houver menos de 9 linhas
                while len(values) < 24:
                    values.append(["", "", ""])  # Preencher com strings vazias e zero (float) 
                result_write_values, log_content = self.write_values(access_token, sheet_name, range_address, values, log_content)  
                if result_write_values > 0:
                    log_entry.save_log(
                        user=user,  # Uma instância do usuário que fez a ação
                        log=log_content,
                        resultado=False,
                        client_id=client_id,  # ID do cliente relacionado
                        date_start=date_start,
                        date_end=date_end
                    )                 
                    return Response({'message': 'Ocorreu um erro ao exportar o dashboard. Tente novamente.'}, status=status.HTTP_400_BAD_REQUEST)                                                            
                ##############
                # Tabela CustosAdicionais 
                sheet_name = "CustosAdicionais"  
                range_address = "A2:B16" 
                # Montar o array values
                values = []

                for item in servicos_adicionais_pagos:
                    values.append([
                        item["name"],
                        float(item["Valor"])
                    ])

                # Preenche com valores padrão se houver menos de 9 linhas
                while len(values) < 15:
                    values.append(["", ""])  # Preencher com strings vazias e zero (float) 
                result_write_values, log_content = self.write_values(access_token, sheet_name, range_address, values, log_content)
                if result_write_values > 0:
                    log_entry.save_log(
                        user=user,  # Uma instância do usuário que fez a ação
                        log=log_content,
                        resultado=False,
                        client_id=client_id,  # ID do cliente relacionado
                        date_start=date_start,
                        date_end=date_end
                    )                 
                    return Response({'message': 'Ocorreu um erro ao exportar o dashboard. Tente novamente.'}, status=status.HTTP_400_BAD_REQUEST)                                                                        
            else:
                current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                log_content += f'[{current_time}] - Erro ao obter token para acessar o drive \n'
                log_entry = LogExport()
                log_entry.save_log(
                    user=user,  # Uma instância do usuário que fez a ação
                    log=log_content,
                    resultado=False,
                    client_id=client_id,  # ID do cliente relacionado
                    date_start=date_start,
                    date_end=date_end
                )            

                return JsonResponse({"error": f"Ocorreu um erro ao gerar o dashboard. Tente novamente"}, status=status.HTTP_400_BAD_REQUEST)
                
            access_token = result['access_token']
            resposta_download = self.download_excel(request, access_token)
            response_data = json.loads(resposta_download.content)        

            if resposta_download.status_code == 200:
                
                download_url = response_data['download_url']
                
                instance = serializer.instance  # Acesse a instância salva
                instance.download_url = download_url  # Atualize o campo
                instance.save()  # Salve novamente para persistir a atualização            
                
                current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                log_content += f'[{current_time}] - Exportação gerada com successo \n'
                
                log_entry = LogExport()
                log_entry.save_log(
                    user=user,  # Uma instância do usuário que fez a ação
                    log=log_content,
                    resultado=True,
                    client_id=client_id,  # ID do cliente relacionado
                    date_start=date_start,
                    date_end=date_end
                ) 
            else:
                error_message = response_data['error']
                current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                log_content += f'[{current_time}] - Erro ao fazer download do dashboard:{error_message} \n'
                log_entry = LogExport()
                log_entry.save_log(
                    user=user,  # Uma instância do usuário que fez a ação
                    log=log_content,
                    resultado=False,
                    client_id=client_id,  # ID do cliente relacionado
                    date_start=date_start,
                    date_end=date_end
                )                   
                
            return resposta_download
        except Exception as e:
            # Registrar o erro
            logger.error(f'Erro ao criar o link compartilhado: {str(e)}')
            return Response({'error': 'Ocorreu um erro ao criar o link compartilhado.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)        

# class SharedLinkCreateView(generics.CreateAPIView):
#     queryset = SharedLink.objects.all()
#     serializer_class = SharedLinkSerializer
#     permission_classes = [IsAuthenticated]

#     def perform_create(self, serializer):
#         # Gera um código único para o code_url
#         unique_code = str(uuid.uuid4())  # Gera um UUID como código único
#         serializer.save(created_by=self.request.user, code_url=unique_code)

#     def create(self, request, *args, **kwargs):
#         # Chama o método de criação padrão
#         serializer = self.get_serializer(data=request.data)
#         serializer.is_valid(raise_exception=True)
#         self.perform_create(serializer)

#         # Retorna a resposta com o code_url
#         return Response({'code_url': serializer.instance.code_url}, status=status.HTTP_201_CREATED)

class ReceivedDataView(APIView):
    
    # Exigir que o usuário esteja autenticado
    permission_classes = [IsAuthenticated]     

    # Helper para formatar o valor bruto
    def format_mil(self, value):
        return f"{value / 1000:,.2f} Mil".replace(",", "X").replace(".", ",").replace("X", ".")

    # Método para criar informações agregadas
    def create_info_numbers(self, query, tipo, formatado):
        valor_bruto_sum = query['valor_bruto_sum'] or 0
        valor_taxa_sum = query['valor_taxa_sum'] or 0
        value = float(valor_taxa_sum / valor_bruto_sum) * 100 if valor_bruto_sum > 0 else 0
        if formatado:
            value = f"{value:.2f}%"  # Formata o valor como percentual

        # Dicionário de títulos e cores
        vetor = {
            'false': {
                'venda': "Venda Sem Antecipação",
                'despesa': "Despesa Sem Antecipação",
                'taxa': "Taxa Sem Antecipação",
                'color': "#7a4019"
            },
            'true': {
                'venda': "Venda Com Antecipação",
                'despesa': "Despesa Com Antecipação",
                'taxa': "Taxa Com Antecipação",
                'color': "#54B435"
            },   
            'total':{
                'venda': "Venda Total",
                'despesa': "Despesa Total",
                'taxa': "Taxa Total",
                'color': "#3871cb"                
            }
        }
        
        # Dicionário com as informações de retorno
        info_numbers = {
            "venda": {},
            "despesa": {},
            "taxa": {},
        }        

        # Preenche info_numbers com as informações formatadas
        info_numbers["venda"] = {
            "title": vetor[tipo]['venda'],
            "icon": "credit_card",
            "growth": 33,
            "value": self.format_mil(valor_bruto_sum) if formatado else valor_bruto_sum,
            "color": vetor[tipo]['color']
        }
        
        info_numbers["despesa"] = {
            "title": vetor[tipo]['despesa'],
            "icon": "credit_card",
            "growth": 33,
            "value": self.format_mil(valor_taxa_sum) if formatado else valor_taxa_sum,
            "color": vetor[tipo]['color']
        }
        
        info_numbers["taxa"] = {
            "title": vetor[tipo]['taxa'],
            "icon": "credit_card",
            "growth": 33,
            "value": value,
            "color": vetor[tipo]['color']
        }

        return info_numbers["venda"], info_numbers["despesa"], info_numbers["taxa"]
    
    def get_total_vendas(self, client_id, date_start, date_end):
        # Filtra os registros de acordo com client_id e intervalo de data
        total_vendas = Received.objects.filter(
            client_id=client_id,
            data_pagamento__range=(date_start, date_end),
            valor_liquido__gte=0  # Adiciona a condição para valor_liquido >= 0
        ).count()  # Conta o número de linhas que atendem ao filtro

        return total_vendas  # Retorna o total de vendas
     
    
    def get_modality_numbers(self, client_id, date_start, date_end):
        # Filtra os registros de acordo com client_id e intervalo de data
        queryset = Received.objects.filter(
            client_id=client_id,
            data_pagamento__range=(date_start, date_end)
        ).values('modality').annotate(
            valor_bruto_sum=Sum('valor_bruto'),
            valor_taxa_sum=Sum('valor_taxa')
        )

        # Monta o resultado calculando a porcentagem e obtendo nome e ID da modality
        result = []
        total_taxa_sum = 0
        total_bruto_sum = 0
        for entry in queryset:
            modality_id = entry['modality']
            modality = get_object_or_404(Modality, id=modality_id)
            valor_bruto_sum = entry['valor_bruto_sum']
            valor_taxa_sum = entry['valor_taxa_sum'] or 0  # Caso `valor_taxa` seja `None`

            # Acumula os totais
            total_bruto_sum += valor_bruto_sum
            total_taxa_sum += valor_taxa_sum

            # Calcula a porcentagem (valor_taxa_sum / valor_bruto_sum) se valor_bruto_sum > 0
            value = valor_taxa_sum / valor_bruto_sum if valor_bruto_sum > 0 else 0
            value = float(value * 100) # convertendo para porcentagem
            
            if value > 0:
                result.append({
                    "id": modality.id,
                    "title": modality.name,
                    "venda": valor_bruto_sum,
                    "taxa": valor_taxa_sum,
                    "value": value
                })

        # Calcular a média e adicioná-la ao resultado
        if total_bruto_sum > 0:  # Evitar divisão por zero
            average_value = (total_taxa_sum / total_bruto_sum) * 100  # média em porcentagem
        else:
            average_value = 0  # Caso não haja valores brutos

        # Adiciona o valor médio ao resultado
        result.append({
            "id": 10,
            "title": "Média",
            "venda": total_bruto_sum,
            "taxa": total_taxa_sum,
            "value": average_value
        })
                
        # Serializa o resultado e retorna a resposta JSON
        modality_data = ModalityResultSerializer(result, many=True)  
        
        return modality_data      
    
    def get_info_numbers(self, queryset, client_id, date_start, date_end, formatado=True):
        # Info Numbers
        info_numbers = {
            "semAntecipacaoVenda":{},
            "semAntecipacaoDespesa":{},
            "semAntecipacaoTaxa":{},
            "antecipacaoVenda":{},
            "antecipacaoDespesa":{},
            "antecipacaoTaxa":{},
            "vendaTotal":{},
            "despesaTotal":{},
            "taxaTotal":{}
        }
        # Agregação para quando idt_antecipacao é False
        antecipacao_false = queryset.filter(
            idt_antecipacao=False,
            client_id=client_id,
            data_pagamento__range=(date_start, date_end)            
        ).aggregate(
            valor_bruto_sum=Sum('valor_bruto'),
            valor_taxa_sum=Sum('valor_taxa')
        )
        
        info_numbers["semAntecipacaoVenda"], info_numbers["semAntecipacaoDespesa"], info_numbers["semAntecipacaoTaxa"] = self.create_info_numbers(antecipacao_false, "false", formatado)

        # Agregação para quando idt_antecipacao é True
        antecipacao_true = queryset.filter(
            idt_antecipacao=True,
            client_id=client_id,
            data_pagamento__range=(date_start, date_end)
        ).aggregate(
            valor_bruto_sum=Sum('valor_bruto'),
            valor_taxa_sum=Sum('valor_taxa')
        )
        info_numbers["antecipacaoVenda"], info_numbers["antecipacaoDespesa"], info_numbers["antecipacaoTaxa"] = self.create_info_numbers(antecipacao_true, "true", formatado)
        
        # Soma total sem filtro
        total_received = Received.objects.filter(
            client_id=client_id,
            data_pagamento__range=(date_start, date_end)
        ).aggregate(
            valor_bruto_sum=Sum('valor_bruto'),
            valor_taxa_sum=Sum('valor_taxa')
        )
        
        info_numbers["vendaTotal"], info_numbers["despesaTotal"], info_numbers["taxaTotal"] = self.create_info_numbers(total_received, "total", formatado)
        
        return info_numbers
                
    def get_tipo_cartoes(self, client_id, date_start, date_end):
        # Query que agrupa os dados por `type_card` e produto
        queryset = (
            Received.objects.filter(
                client_id=client_id,
                data_pagamento__range=(date_start, date_end)
            ).values(
                'product__type_card__name',  # Nome do tipo de cartão
                'product__name'              # Nome do produto
            )
            .annotate(
                venda_bruta=Sum('valor_bruto'),  # Soma do valor bruto
                valor_taxa=Sum('valor_taxa')     # Soma do valor taxa
            )
            .order_by('product__type_card__name')  # Ordena por tipo do cartão
        )       
        
        # Estrutura para o JSON de resposta
        tipo_cartoes = {}

        for entry in queryset:
            type_card_name = entry['product__type_card__name']
            product_name = entry['product__name']
            venda_bruta = entry['venda_bruta'] or 0
            valor_taxa = entry['valor_taxa'] or 0
            # Calcula a taxa percentual
            taxa_porcentagem = (valor_taxa / venda_bruta * 100) if venda_bruta else 0

            # Constrói cada entrada para o tipo de cartão
            product_data = {
                "name": product_name,
                "Venda Bruta": venda_bruta,
                "Taxa R$": valor_taxa,
                "Taxa%": round(taxa_porcentagem, 2)
            }

            # Adiciona os produtos agrupados por `type_card.name`
            if type_card_name not in tipo_cartoes:
                tipo_cartoes[type_card_name] = []
            tipo_cartoes[type_card_name].append(product_data)

        # Formata o JSON final
        tipo_cartoes_data = [
            {type_card_name: products} for type_card_name, products in tipo_cartoes.items()
        ]

        # Ajusta a estrutura final para retornar apenas um dicionário
        tipo_cartoes_final = [tipo_cartoes]  # Coloca em uma lista como requerido
 
        return tipo_cartoes_final        
    
    def get_adquirente(self, client_id, date_start, date_end):
        # Query que agrupa os dados por `acquirer`
        acquirer_queryset = (
            Received.objects.filter(
                client_id=client_id,
                data_pagamento__range=(date_start, date_end)
            ).values(
                'adquirente__name'  # Nome do adquirente
            )
            .annotate(
                venda_bruta=Sum('valor_bruto'),  # Soma do valor bruto
                soma_valor_taxa=Sum('valor_taxa'),  # Soma do valor taxa
                taxa_percentual=(Sum('valor_taxa') / Sum('valor_bruto') * 100)  # Taxa percentual
            )
            .order_by('adquirente__name')  # Ordena por nome do adquirente
        )

        # Estrutura para o JSON de resposta
        adquirente_data = []

        for entry in acquirer_queryset:
            name = entry['adquirente__name']
            venda_bruta = entry['venda_bruta'] or 0
            soma_valor_taxa = entry['soma_valor_taxa'] or 0
            taxa_percentual = round(entry['taxa_percentual'], 2) if venda_bruta else 0

            # Cria a entrada para o adquirente
            acquirer_info = {
                "name": name,
                "Venda Bruta": venda_bruta,
                "Soma_de_Valor_Taxa": soma_valor_taxa,
                "Taxa%": taxa_percentual
            }
            adquirente_data.append(acquirer_info)

        return adquirente_data
                    
    def get_servicos_adicionais_pagos(self, client_id, date_start, date_end):
        # Query que agrupa os dados por `observacao`, somando `valor_liquido` onde for menor que 0
        observation_queryset = (
            Received.objects.filter(
                client_id=client_id,
                data_pagamento__range=(date_start, date_end),
                valor_liquido__lt=0  # Filtra para incluir apenas valores líquidos menores que 0
            )
            .annotate(
                servicos_pagos_name=Subquery(
                    ServicosPagos.objects.filter(observacao=OuterRef('observacao')).values('name')[:1]
                )
            )
            .values('servicos_pagos_name')  # Agrupa pelo nome de ServicosPagos
            .annotate(
                total_liquido=Sum('valor_liquido')  # Soma dos valores líquidos
            )
            .order_by('-total_liquido')  # Ordena do maior para o menor
        )

        # Cálculo do valor total
        total_valor_liquido = observation_queryset.aggregate(total=Sum('total_liquido'))['total'] or 0

        # Estrutura para o JSON de resposta
        servicos_adicionais_pagos = [
            {
                "name": "Total",
                "Valor": round(total_valor_liquido * -1, 2)  # Total dos valores líquidos
            }
        ]

        # Adiciona cada entrada para observação ao JSON
        for entry in observation_queryset:
            observacao = entry['servicos_pagos_name'] or "Sem Observação"  # Use "Sem Observação" se for None
            valor_liquido = round(entry['total_liquido'] * -1, 2)  # Arredonda o valor líquido

            # Adiciona a observação ao JSON
            servicos_adicionais_pagos.append({
                "name": observacao,
                "Valor": valor_liquido
            })            

        servicos_adicionais_pagos = sorted(servicos_adicionais_pagos, key=lambda x: x['Valor'], reverse=True)
        
        return servicos_adicionais_pagos
    
    def get_servicos_adicionais_pagos_bruto(self, client_id, date_start, date_end):
        # Query que agrupa os dados por `observacao`, somando `valor_liquido` onde for menor que 0
        observation_queryset = (
            Received.objects.filter(
                client_id=client_id,
                data_pagamento__range=(date_start, date_end),
                valor_liquido__lt=0  # Filtra para incluir apenas valores líquidos menores que 0
            )
            .values('observacao')  # Agrupa pelo nome de ServicosPagos
            .annotate(
                total_liquido=Sum('valor_liquido')  # Soma dos valores líquidos
            )
            .order_by('-total_liquido')  # Ordena do maior para o menor
        )
        
        queryset_sp = ServicosPagos.objects.all()

        # Cálculo do valor total
        total_valor_liquido = observation_queryset.aggregate(total=Sum('total_liquido'))['total'] or 0

        # Estrutura para o JSON de resposta
        servicos_adicionais_pagos = [
            {
                "name": "Total",
                "name_value": "Total",
                "Valor": round(total_valor_liquido * -1, 2)  # Total dos valores líquidos
            }
        ]

        # Converte o queryset em um dicionário com observacao como chave e name como valor
        observacao_dict = {sp.observacao: sp.name for sp in queryset_sp}

        # Adiciona cada entrada para observação ao JSON
        for entry in observation_queryset:
            observacao = entry['observacao'] or "Sem Observação"  # Use "Sem Observação" se for None
            valor_liquido = round(entry['total_liquido'] * -1, 2)  # Arredonda o valor líquido
            # Verifica se o valor de entry['observacao'] está no dicionário e retorna o valor de name, ou uma string vazia se não estiver
            name_value = observacao_dict.get(entry['observacao'], '')            

            # Adiciona a observação ao JSON
            servicos_adicionais_pagos.append({
                "name": observacao,
                "name_value": name_value,
                "Valor": valor_liquido
            })            

        servicos_adicionais_pagos = sorted(servicos_adicionais_pagos, key=lambda x: x['Valor'], reverse=True)
        
        return servicos_adicionais_pagos    
                
    
    def get(self, request):
        client_id = request.query_params.get('client_id')
        date_start = request.query_params.get('date_start')
        date_end = request.query_params.get('date_end')

        # Verifica se todos os parâmetros obrigatórios foram fornecidos
        if not (client_id and date_start and date_end):
            return Response({"error": "Parâmetros insuficientes."}, status=status.HTTP_400_BAD_REQUEST)

        # Converte strings de data para objetos datetime
        try:
            date_start = datetime.strptime(date_start, "%Y-%m-%d").date()
            date_end = datetime.strptime(date_end, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Formato de data inválido. Use AAAA-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

        # Filtra os registros de acordo com client_id e intervalo de data
        queryset = Received.objects.filter(
            client_id=client_id,
            data_pagamento__range=(date_start, date_end)
        ).values('modality').annotate(
            valor_bruto_sum=Sum('valor_bruto'),
            valor_taxa_sum=Sum('valor_taxa')
        )

        # # Serializa o resultado e retorna a resposta JSON
        modality_data = self.get_modality_numbers(client_id, date_start, date_end)
        info_numbers = self.get_info_numbers(queryset, client_id, date_start, date_end)
        tipo_cartoes_final = self.get_tipo_cartoes(client_id, date_start, date_end)
        adquirente_data = self.get_adquirente(client_id, date_start, date_end)
        servicos_adicionais_pagos = self.get_servicos_adicionais_pagos(client_id, date_start, date_end)
        
        # Primeiro, calcula o valor bruto total para o client_id e no intervalo de datas
        total_venda_bruta = Received.objects.filter(
            client_id=client_id,
            data_pagamento__range=(date_start, date_end)
        ).aggregate(total=Sum('valor_bruto'))['total'] or 0

        # Em seguida, calcula a venda bruta por tipo de cartão
        venda_por_tipo_cartao_queryset = (
            Received.objects.filter(
                client_id=client_id,
                data_pagamento__range=(date_start, date_end),
                product__type_card__id__isnull=False  # Filtra para garantir que o ID do tipo de cartão não seja nulo
            )
            .values('product__type_card__id', 'product__type_card__name')  # Agrupa pelo ID e título do type_card
            .annotate(
                venda_bruta=Sum('valor_bruto')  # Soma das vendas brutas
            )
        )

        # Constrói a lista de resultados para vendaPorTipoCartao
        venda_por_tipo_cartao = []
        for entry in venda_por_tipo_cartao_queryset:
            tipo_cartao_id = entry['product__type_card__id']
            tipo_cartao_title = entry['product__type_card__name']
            venda_bruta = entry['venda_bruta']

            # Calcula a porcentagem em relação à venda bruta total
            if total_venda_bruta > 0:
                porcentagem = (venda_bruta / total_venda_bruta) * 100
            else:
                porcentagem = 0  # Se não houver vendas, a porcentagem é 0

            # Adiciona o dicionário ao resultado, formatando o valor para 2 casas decimais
            venda_por_tipo_cartao.append({
                "id": tipo_cartao_id,
                "title": tipo_cartao_title,
                "value": round(porcentagem, 2) 
            })       

        analytics_data = {
            "TaxaEfetiva": modality_data.data,
            "infoNumbers": info_numbers,
            "tipoCartoes": tipo_cartoes_final,
            "adquirente": adquirente_data,
            "servicosAdicionaisPagos": servicos_adicionais_pagos,
            "vendaPorTipoCartao": venda_por_tipo_cartao
        }        
        
        return Response(analytics_data, status=status.HTTP_200_OK)

class SharedLinkDashboardView(APIView): 
    permission_classes = [AllowAny]
         
    def get(self, request):
        # Obter parâmetros da URL
        code = request.query_params.get('code')
        # Validar parâmetros obrigatórios
        if not code:
            return Response({"error": "Código não encontrado."},
                            status=status.HTTP_400_BAD_REQUEST)        
        
        shared = get_object_or_404(SharedLink, code_url=code)
        client_id = shared.client.pk
        fantasy_name = shared.client.fantasy_name
        info_adicional = shared.info
        date_start = shared.date_start
        date_end = shared.date_end
        # Formatar as datas no padrão desejado
        date_start_formatted = date_start.strftime("%d/%m/%Y")
        date_end_formatted = date_end.strftime("%d/%m/%Y")

        # Criar o texto com a quebra de linha
        periodo_apuracao = f"{date_start_formatted} - {date_end_formatted}"
        
        # Filtra os registros de acordo com client_id e intervalo de data
        queryset = Received.objects.filter(
            client_id=client_id,
            data_pagamento__range=(date_start, date_end)
        ).values('modality').annotate(
            valor_bruto_sum=Sum('valor_bruto'),
            valor_taxa_sum=Sum('valor_taxa')
        )          
        
        # Obtenha o cliente pelo client_id e retorne o fantasy_name
        client = get_object_or_404(Clients, id=client_id)
        fantasy_name = client.fantasy_name        
        
        # Serializa o resultado e retorna a resposta JSON
        received_data_view = ReceivedDataView()
        modality_data = received_data_view.get_modality_numbers(client_id, date_start, date_end)
        info_numbers = received_data_view.get_info_numbers(queryset, client_id, date_start, date_end)
        tipo_cartoes_final = received_data_view.get_tipo_cartoes(client_id, date_start, date_end)
        adquirente_data = received_data_view.get_adquirente(client_id, date_start, date_end)
        servicos_adicionais_pagos = received_data_view.get_servicos_adicionais_pagos(client_id, date_start, date_end)          
        
        # Primeiro, calcula o valor bruto total para o client_id e no intervalo de datas
        total_venda_bruta = Received.objects.filter(
            client_id=client_id,
            data_pagamento__range=(date_start, date_end)
        ).aggregate(total=Sum('valor_bruto'))['total'] or 0

        # Em seguida, calcula a venda bruta por tipo de cartão
        venda_por_tipo_cartao_queryset = (
            Received.objects.filter(
                client_id=client_id,
                data_pagamento__range=(date_start, date_end),
                product__type_card__id__isnull=False  # Filtra para garantir que o ID do tipo de cartão não seja nulo
            )
            .values('product__type_card__id', 'product__type_card__name')  # Agrupa pelo ID e título do type_card
            .annotate(
                venda_bruta=Sum('valor_bruto')  # Soma das vendas brutas
            )
        )

        # Constrói a lista de resultados para vendaPorTipoCartao
        venda_por_tipo_cartao = []
        for entry in venda_por_tipo_cartao_queryset:
            tipo_cartao_id = entry['product__type_card__id']
            tipo_cartao_title = entry['product__type_card__name']
            venda_bruta = entry['venda_bruta']

            # Calcula a porcentagem em relação à venda bruta total
            if total_venda_bruta > 0:
                porcentagem = (venda_bruta / total_venda_bruta) * 100
            else:
                porcentagem = 0  # Se não houver vendas, a porcentagem é 0

            # Adiciona o dicionário ao resultado, formatando o valor para 2 casas decimais
            venda_por_tipo_cartao.append({
                "id": tipo_cartao_id,
                "title": tipo_cartao_title,
                "value": round(porcentagem, 2) 
            })       

        analytics_data = {
            "NomeFantasia": fantasy_name,
            "InfoAdicional": info_adicional,
            "Periodo": periodo_apuracao,
            "TaxaEfetiva": modality_data.data,
            "infoNumbers": info_numbers,
            "tipoCartoes": tipo_cartoes_final,
            "adquirente": adquirente_data,
            "servicosAdicionaisPagos": servicos_adicionais_pagos,
            "vendaPorTipoCartao": venda_por_tipo_cartao
        }        
        
        return Response(analytics_data, status=status.HTTP_200_OK)        

class ExportDashboardView(generics.CreateAPIView):
    queryset = SharedLink.objects.all()
    serializer_class = SharedLinkSerializer
    permission_classes = [IsAuthenticated]
    
    def perform_create(self, serializer):
        # Gera um código único para o code_url
        unique_code = str(uuid.uuid4())  # Gera um UUID como código único
        serializer.save(created_by=self.request.user, code_url=unique_code)

    def create(self, request, *args, **kwargs):
        try:
            # Chama o método de criação padrão
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            code_url = serializer.instance.code_url
            pot_econ_acao_01 = serializer.instance.info_action_01
            pot_econ_acao_02 = serializer.instance.info_action_02
            pot_econ_acao_03 = serializer.instance.info_action_03
            pot_econ_acao_04 = serializer.instance.info_action_04
            pot_econ_01 = serializer.instance.action_01
            pot_econ_02 = serializer.instance.action_02
            pot_econ_03 = serializer.instance.action_03
            pot_econ_04 = serializer.instance.action_04
                    
            # Obter o objeto criado para usar nas operações a seguir
            user = request.user 
            shared = serializer.instance
            client_id = shared.client.pk
            fantasy_name = shared.client.fantasy_name
            date_start = shared.date_start
            date_end = shared.date_end
            
            # Montar o log
            current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_content = f'[{current_time}] - Iniciando a exportação do cliente: {fantasy_name} \n'
            
            # Formatar as datas no padrão desejado
            date_start_formatted = date_start.strftime("%d/%m/%Y")
            date_end_formatted = date_end.strftime("%d/%m/%Y")
            periodo_apuracao = f"Período de Apuração \n{date_start_formatted} - {date_end_formatted}"

            # Filtrar registros de acordo com o client_id e intervalo de data
            queryset = Received.objects.filter(
                client_id=client_id,
                data_pagamento__range=(date_start, date_end)
            ).values('modality').annotate(
                valor_bruto_sum=Sum('valor_bruto'),
                valor_taxa_sum=Sum('valor_taxa')
            )          

            # Obter informações adicionais
            client = get_object_or_404(Clients, id=client_id)
            fantasy_name = client.fantasy_name        

            received_data_view = ReceivedDataView()
            modality_data = received_data_view.get_modality_numbers(client_id, date_start, date_end)
            info_numbers = received_data_view.get_info_numbers(queryset, client_id, date_start, date_end, formatado=False)
            tipo_cartoes_final = received_data_view.get_tipo_cartoes(client_id, date_start, date_end)
            adquirente_data = received_data_view.get_adquirente(client_id, date_start, date_end)
            servicos_adicionais_pagos = received_data_view.get_servicos_adicionais_pagos_bruto(client_id, date_start, date_end)          
            quantidade_total_vendas = received_data_view.get_total_vendas(client_id, date_start, date_end)          
            venda_total = info_numbers["vendaTotal"]["value"]
                       
            # CRIANDO O EXCEL
            # Criar um novo Workbook
            wb = Workbook()
            
            # Selecionar a planilha ativa e renomeá-la
            ws1 = wb.active
            ws1.title = "DRIVE DE VENDAS"
            percentage_style = NamedStyle(name="percentage_style", number_format='0.00%')
            current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_content += f'[{current_time}] - Iniciando a planilha: DRIVE DE VENDAS \n'

            # Adicionar cabeçalho na primeira linha
            headers = ["Antecipado", "MODALIDADE", "Bruto", "Taxa R$", "Taxa %"]
            for col_num, header in enumerate(headers, 1):  # Enumerate começa de 1 para coincidir com colunas do Excel
                ws1.cell(row=1, column=col_num, value=header) 
                
            ws1["A2"] = "falso"
            ws1["B2"] = "VENDA SEM ANTECIPAÇÃO"
            ws1["C2"] = info_numbers["semAntecipacaoVenda"]["value"]
            ws1["D2"] = info_numbers["semAntecipacaoDespesa"]["value"]
            ws1["E2"] = info_numbers["semAntecipacaoTaxa"]["value"]/100
            ws1["E2"].style = percentage_style
            
            ws1["A3"] = "verdadeiro"
            ws1["B3"] = "VENDA COM ANTECIPAÇÃO"
            ws1["C3"] = info_numbers["antecipacaoVenda"]["value"]
            ws1["D3"] = info_numbers["antecipacaoDespesa"]["value"]
            ws1["E3"] = info_numbers["antecipacaoTaxa"]["value"]/100
            ws1["E3"].style = percentage_style
            
            ws1["A4"] = "VENDA TOTAL"
            ws1["B4"] = "VENDA TOTAL"
            ws1["C4"] = info_numbers["vendaTotal"]["value"]
            ws1["D4"] = info_numbers["despesaTotal"]["value"]
            ws1["E4"] = info_numbers["taxaTotal"]["value"]/100
            ws1["E4"].style = percentage_style
                       
            # Adicionar uma nova planilha
            current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_content += f'[{current_time}] - Iniciando a planilha: VENDA TAXA MEDIA PART POR ADIQU \n'
                        
            ws2 = wb.create_sheet(title="VENDA TAXA MEDIA PART POR ADIQU")            
            # Adicionar cabeçalho na primeira linha
            headers = ["ADQUIRENTE", "VENDA BRUTA", "TAXA R$", "TAXA %", "% | Venda"]
            for col_num, header in enumerate(headers, 1):  # Enumerate começa de 1 para coincidir com colunas do Excel
                ws2.cell(row=1, column=col_num, value=header) 
            
            values = []
            
            for item in adquirente_data:
                values.append([
                    item["name"],
                    float(item["Venda Bruta"]),
                    float(item["Soma_de_Valor_Taxa"]),
                    float(item["Taxa%"])/100,
                    float(item["Venda Bruta"] / venda_total) if float(item["Venda Bruta"]) > 0 else 0        
                ])  
                
            for row_num, row_data in enumerate(values, 2):  # Começa da linha 2
                for col_num, cell_value in enumerate(row_data, 1):
                    cell = ws2.cell(row=row_num, column=col_num, value=cell_value)

                    # Verifica se a coluna é D (4) ou E (5)
                    if col_num in [4, 5]:  # Colunas D e E
                        cell.style = percentage_style  # Aplica o estilo de porcentagem                    
            
            # Adicionar uma nova planilha
            current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_content += f'[{current_time}] - Iniciando a planilha: SERVIÇOS ADICONAIS PAGOS R$ \n'            
            
            ws3 = wb.create_sheet(title="SERVIÇOS ADICONAIS PAGOS R$")
            # Adicionar cabeçalho na primeira linha
            headers = ["REFERÊNCIA - FORMULA", "SERVIÇOS COBRADOS", "VALOR R$"]
            for col_num, header in enumerate(headers, 1):  # Enumerate começa de 1 para coincidir com colunas do Excel
                ws3.cell(row=1, column=col_num, value=header) 
                        
            values = []
            for item in servicos_adicionais_pagos:
                values.append([
                    item["name"],
                    item["name_value"],
                    float(item["Valor"])
                ])
                
            for row_num, row_data in enumerate(values, 2):  # Começa da linha 2
                for col_num, cell_value in enumerate(row_data, 1):
                    ws3.cell(row=row_num, column=col_num, value=cell_value)                                                

            # Adicionar uma nova planilha
            current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_content += f'[{current_time}] - Iniciando a planilha: TAXA EFETIVA \n'                        
            ws4 = wb.create_sheet(title="TAXA EFETIVA")
            # Adicionar cabeçalho na primeira linha
            headers = ["MODALIDADE", "GRAFICO", "Bruto", "Taxa R$", "Taxa %"]
            for col_num, header in enumerate(headers, 1):  # Enumerate começa de 1 para coincidir com colunas do Excel
                ws4.cell(row=1, column=col_num, value=header) 
            
            values = []   
            # Adiciona os pares [title, value] ao array values
            for item in modality_data.data:
                values.append([
                    item["title"].upper(),
                    item["title"].title(),
                    float(item["venda"]),
                    float(item["taxa"]),
                    float(item["value"])/100
                ])
                                
            for row_num, row_data in enumerate(values, 2):  # Começa da linha 2
                for col_num, cell_value in enumerate(row_data, 1):
                    cell = ws4.cell(row=row_num, column=col_num, value=cell_value)

                    # Verifica se a coluna é E (5)
                    if col_num in [5]:  # E
                        cell.style = percentage_style  # Aplica o estilo de porcentagem                    
            
            
            # Adicionar uma nova planilha
            current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_content += f'[{current_time}] - Iniciando a planilha: GRAFICO DE ABERTURA BANDEIRAS \n'                        
            ws5 = wb.create_sheet(title="GRAFICO DE ABERTURA BANDEIRAS")
            # Adicionar cabeçalho na primeira linha
            			
            headers = ["BANDEIRA", "Venda Bruta", "Taxa R$", "Taxa %"]
            for col_num, header in enumerate(headers, 1):  # Enumerate começa de 1 para coincidir com colunas do Excel
                ws5.cell(row=1, column=col_num, value=header) 
            
            values = []   
            debito_items = tipo_cartoes_final[0]["Debito"]
            # Montar o array values com os itens de "Debito"
            for item in debito_items:
                values.append([
                    item["name"].upper(),
                    float(item["Venda Bruta"]),
                    float(item["Taxa R$"]),
                    float(item["Taxa%"])/100
                ])
            credito_items = tipo_cartoes_final[0]["Credito"]
            # Montar o array values com os itens de "Credito"
            for item in credito_items:
                values.append([
                    item["name"].upper(),
                    float(item["Venda Bruta"]),
                    float(item["Taxa R$"]),
                    float(item["Taxa%"])/100
                ])  
            voucher_items = tipo_cartoes_final[0]["Voucher"] 
            # Montar o array values com os itens de "Credito"
            for item in voucher_items:
                values.append([
                    item["name"].upper(),
                    float(item["Venda Bruta"]),
                    float(item["Taxa R$"]),
                    float(item["Taxa%"])/100
                ])                      
                                
            for row_num, row_data in enumerate(values, 2):  # Começa da linha 2
                for col_num, cell_value in enumerate(row_data, 1):
                    cell = ws5.cell(row=row_num, column=col_num, value=cell_value)

                    # Verifica se a coluna é D (4)
                    if col_num in [4]:  # D
                        cell.style = percentage_style  # Aplica o estilo de porcentagem                    
                            

            # Adicionar uma nova planilha
            current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_content += f'[{current_time}] - Iniciando a planilha: POTENCIAL DE ECONOMIA \n'
            
            ws6 = wb.create_sheet(title="POTENCIAL DE ECONOMIA") 
            # Adicionar cabeçalho na primeira linha
            headers = ["AÇÕES", "ECONOMIA - R$"]
            for col_num, header in enumerate(headers, 1):  # Enumerate começa de 1 para coincidir com colunas do Excel
                ws6.cell(row=1, column=col_num, value=header)             
            
            values = [
                [pot_econ_acao_01, pot_econ_01],
                [pot_econ_acao_02, pot_econ_02],
                [pot_econ_acao_03, pot_econ_03],
                [pot_econ_acao_04, pot_econ_04]
            ] 
            
            for row_num, row_data in enumerate(values, 2):  # Começa da linha 2
                for col_num, cell_value in enumerate(row_data, 1):
                    ws6.cell(row=row_num, column=col_num, value=cell_value)                                                                 

            # Adicionar uma nova planilha
            current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_content += f'[{current_time}] - Iniciando a planilha: QTDE DE VENDAS \n'            
            
            ws7 = wb.create_sheet(title="QTDE DE VENDAS") 
            ws7["A1"] = "QTDE VENDAS"
            ws7["A2"] = quantidade_total_vendas
            
            # Adicionar uma nova planilha
            current_time = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            log_content += f'[{current_time}] - Iniciando a planilha: IMPACTO ANUAL ANTECIPAÇÃO \n'                                     
            
            ws8 = wb.create_sheet(title="IMPACTO ANUAL ANTECIPAÇÃO") 
            valor_numero = info_numbers["antecipacaoDespesa"]["value"]

            # Multiplicar por 12
            impacto_anual_antecipacao = valor_numero * 12              
            
            ws8["A1"] = "IMPACTO ANUAL ANTECIPAÇÃO (SEM TOMAR AÇÃO)"
            ws8["A2"] = impacto_anual_antecipacao 

            # Salvar o workbook em um buffer de memória
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            log_entry = LogExport()
            log_entry.save_log(
                user=user,  # Uma instância do usuário que fez a ação
                log=log_content,
                resultado=True,
                client_id=client_id,  # ID do cliente relacionado
                date_start=date_start,
                date_end=date_end
            )             

            # Configurar a resposta para o download do arquivo
            fantasy_name_cleaned = re.sub(r'[^A-Za-z0-9]', '', fantasy_name.title().replace(" ", ""))
            
            # Converter datas para o formato DDMMYYYY
            date_start_str = date_start.strftime('%d%m%Y')
            date_end_str = date_end.strftime('%d%m%Y')
            
            # Gerar o nome do arquivo final
            filename = f"{fantasy_name_cleaned}_{date_start_str}_{date_end_str}.xlsx"            

            file_path = os.path.join(settings.MEDIA_ROOT, 'exported_dashboards', filename)

            # Certificar que o diretório existe
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            # Salvar o conteúdo do arquivo localmente
            with open(file_path, 'wb') as file:
                file.write(buffer.getvalue())

            # Gerar URL para o arquivo
            file_url = f"{settings.MEDIA_URL}exported_dashboards/{filename}"
            file_url = request.build_absolute_uri(file_url)

            instance = serializer.instance  # Acesse a instância salva
            instance.download_url = file_url  # Atualize o campo
            instance.save()  # Salve novamente para persistir a atualização 
            
            # response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            # response['Content-Disposition'] = 'attachment; filename={filename}'

            # return response
            return JsonResponse({"download_url": file_url}, status=status.HTTP_200_OK)            
        except Exception as e:
            # Registrar o erro
            logger.error(f'Erro ao criar o link compartilhado: {str(e)}')
            return Response({'error': 'Ocorreu um erro ao criar o link compartilhado.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)        




class ExportDashboardViewFirstVersion(APIView):
    # Exigir que o usuário esteja autenticado
    permission_classes = [IsAuthenticated]         
    
    def get(self, request):
        # Obter parâmetros da URL
        client_id = request.query_params.get('client_id')
        date_start = request.query_params.get('date_start')
        date_end = request.query_params.get('date_end')
        
        # Validar parâmetros obrigatórios
        if not client_id or not date_start or not date_end:
            return Response({"error": "Parâmetros 'client_id', 'date_start' e 'date_end' são obrigatórios."},
                            status=status.HTTP_400_BAD_REQUEST)
            
        # Converte strings de data para objetos datetime
        try:
            date_start = datetime.strptime(date_start, "%Y-%m-%d").date()
            date_end = datetime.strptime(date_end, "%Y-%m-%d").date()
        except ValueError:
            return Response({"error": "Formato de data inválido. Use AAAA-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

        # Filtra os registros de acordo com client_id e intervalo de data
        queryset = Received.objects.filter(
            client_id=client_id,
            data_pagamento__range=(date_start, date_end)
        ).values('modality').annotate(
            valor_bruto_sum=Sum('valor_bruto'),
            valor_taxa_sum=Sum('valor_taxa')
        )          
        
        # Obtenha o cliente pelo client_id e retorne o fantasy_name
        client = get_object_or_404(Clients, id=client_id)
        fantasy_name = client.fantasy_name        
        
        # Serializa o resultado e retorna a resposta JSON
        received_data_view = ReceivedDataView()
        # modality_data = received_data_view.get_modality_numbers(client_id, date_start, date_end)
        # info_numbers = received_data_view.get_info_numbers(queryset, client_id, date_start, date_end)
        # tipo_cartoes_final = received_data_view.get_tipo_cartoes(client_id, date_start, date_end)
        # adquirente_data = received_data_view.get_adquirente(client_id, date_start, date_end)
        # servicos_adicionais_pagos = received_data_view.get_servicos_adicionais_pagos(client_id, date_start, date_end)          

        # Caminho para o modelo do Excel em dashboards/excel_model/
        model_path = os.path.join(settings.BASE_DIR, 'dashboards', 'excel_model', 'dashboard_model.xlsx')
        
        # Carregar o modelo do Excel
        # workbook = load_workbook(model_path)
        workbook = openpyxl.load_workbook(model_path, read_only=False)

        sheet = workbook['Info']  # Use o nome exato da aba
        
        # # Colocando os dados básicos
        formatted_text = f'Período de Apuração {date_start.strftime("%d/%m/%Y")} - {date_end.strftime("%d/%m/%Y")}'
        sheet['B1'] = f'{fantasy_name}'
        sheet['B2'] = f'{formatted_text}'
        
        # Caminho para salvar o arquivo atualizado na pasta `media`
        # output_filename = f"dashboard_{client_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        # output_path = os.path.join(settings.MEDIA_ROOT, 'exported_dashboards', output_filename)

        # Criar a pasta se não existir
        # os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Salvar o workbook atualizado
        # workbook.save(output_path)
        workbook.save(model_path)
        
        return Response({"error": "Formato de data inválido. Use AAAA-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

        # Retornar o arquivo atualizado para download
        # with open(output_path, 'rb') as file:
        #     response = HttpResponse(
        #         file.read(),
        #         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        #     )
        #     response['Content-Disposition'] = f'attachment; filename={output_filename}'
        #     return response



def dashboard_view(request):
    analytics_data = {
        "TaxaEfetiva":[
            {
                "id": 1,
                "title": "Benefício",
                "value": 6.13 
            },
            {
                "id": 2,
                "title": "Crédito a Vista",
                "value": 2.55 
            },
            {
                "id": 3,
                "title": "Débito",
                "value": 1.32 
            },
            {
                "id": 4,
                "title": "Pix",
                "value": 0.48 
            },                                 
        ],
        "infoNumbers": {
            "semAntecipacaoVenda": {
                "title": "Venda Sem Antecipação",
                "icon": "credit_card",
                "growth": 33,
                "value": "1.580.70 Mil",
                "color": "#7a4019",
            },
            "semAntecipacaoDespesa": {
                "title": "Despesa Sem Antecipação",
                "icon": "credit_card",
                "growth": 33,
                "value": "39.77 Mil",
                "color": "#7a4019",
            },
            "semAntecipacaoTaxa": {
                "title": "Taxa Sem Antecipação",
                "icon": "credit_card",
                "growth": 33,
                "value": "2,52%",
                "color": "#7a4019",
            },
            "antecipacaoVenda": {
                "title": "Venda Com Antecipação",
                "icon": "credit_card",
                "growth": 33,
                "value": "255.31 Mil",
                "color": "#54B435",
            },
            "antecipacaoDespesa": {
                "title": "Despesa Com Antecipação",
                "icon": "credit_card",
                "growth": 33,
                "value": "15.35 Mil",
                "color": "#54B435",
            },
            "antecipacaoTaxa": {
                "title": "Taxa com Antecipação",
                "icon": "credit_card",
                "growth": 33,
                "value": "6,00%",
                "color": "#54B435",
            },   
            "vendaTotal": {
                "title": "Venda Total",
                "icon": "credit_card",
                "growth": 33,
                "value": "1.836.01 Mil",
                "color": "#3871cb",
            },
            "despesaTotal": {
                "title": "Despesa Total",
                "icon": "credit_card",
                "growth": 33,
                "value": "55.09 Mil",
                "color": "#3871cb",
            },
            "taxaTotal": {
                "title": "Taxa Total Cartão",
                "icon": "credit_card",
                "growth": 33,
                "value": "3,00%",
                "color": "#3871cb",
            }
        },
        "tipoCartoes":[
            {
                "Credito": [
                    {
                        "name": "VISA CREDITO A VISTA",
                        "Venda Bruta": 177865.31,
                        "Taxa R$": 4400.62,
                        "Taxa%": 2.47
                    },
                    {
                        "name": "MASTER CREDITO",
                        "Venda Bruta": 306684.54,
                        "Taxa R$": 7293.37,
                        "Taxa%": 2.38
                    },
                    {
                        "name": "PRÉ-PAGO ELO CRÉDITO",
                        "Venda Bruta": 8936.62,
                        "Taxa R$": 304.21,
                        "Taxa%": 3.40
                    },
                    {
                        "name": "PRÉ-PAGO VISA CREDITO",
                        "Venda Bruta": 7641.05,
                        "Taxa R$": 147.71,
                        "Taxa%": 1.93
                    },
                    {
                        "name": "ELO CREDITO",
                        "Venda Bruta": 45188.95,
                        "Taxa R$": 1625.29,
                        "Taxa%": 3.60
                    },
                    {
                        "name": "PRÉ-PAGO MASTER CRÉDITO",
                        "Venda Bruta": 2143.96,
                        "Taxa R$": 74.95,
                        "Taxa%": 3.50
                    },
                    {
                        "name": "TRICARD CRÉDITO À VISTA",
                        "Venda Bruta": 16081.45,
                        "Taxa R$": 514.58,
                        "Taxa%": 3.20
                    },
                    {
                        "name": "AMEX CREDITO A VISTA",
                        "Venda Bruta": 2111.17,
                        "Taxa R$": 72.14,
                        "Taxa%": 3.42
                    },
                    {
                        "name": "HIPERCARD CREDITO A VISTA",
                        "Venda Bruta": 156.60,
                        "Taxa R$": 8.39,
                        "Taxa%": 5.36
                    }
                ],
                "Debito": [
                    {
                        "name": "MAESTRO",
                        "Venda Bruta": 282050.31,
                        "Taxa R$": 2820.43,
                        "Taxa%": 1.00
                    },
                    {
                        "name": "PRÉ-PAGO VISA DÉBITO",
                        "Venda Bruta": 37503.72,
                        "Taxa R$": 457.79,
                        "Taxa%": 1.22
                    },
                    {
                        "name": "VISA ELECTRON DEBITO A VISTA",
                        "Venda Bruta": 277700.84,
                        "Taxa R$": 3230.42,
                        "Taxa%": 1.16
                    },
                    {
                        "name": "ELO DEBITO A VISTA",
                        "Venda Bruta": 151115.83,
                        "Taxa R$": 3107.61,
                        "Taxa%": 2.06
                    },
                    {
                        "name": "PRÉ-PAGO MASTER DÉBITO",
                        "Venda Bruta": 22834.11,
                        "Taxa R$": 552.53,
                        "Taxa%": 2.42
                    },
                    {
                        "name": "PRÉ-PAGO ELO DÉBITO",
                        "Venda Bruta": 224.91,
                        "Taxa R$": 5.62,
                        "Taxa%": 2.50
                    },
                    {
                        "name": "PIX",
                        "Venda Bruta": 35.60,
                        "Taxa R$": 0.17,
                        "Taxa%": 0.48
                    }
                ],
                "Voucher": [
                    {
                        "name": "VALECARD VOUCHER",
                        "Venda Bruta": 13688.32,
                        "Taxa R$": 889.79,
                        "Taxa%": 6.50
                    },                    
                    {
                        "name": "ALELO ALIMENTAÇÃO",
                        "Venda Bruta": 128732.07,
                        "Taxa R$": 8367.75,
                        "Taxa%": 6.50
                    },
                    {
                        "name": "ALELO REFEIÇÃO",
                        "Venda Bruta": 22144.82,
                        "Taxa R$": 1439.39,
                        "Taxa%": 6.50
                    },
                    {
                        "name": "SODEXO PREMIUM PASS",
                        "Venda Bruta": 547.87,
                        "Taxa R$": 32.88,
                        "Taxa%": 6.00
                    },
                    {
                        "name": "SODEXO REFEICAO",
                        "Venda Bruta": 2814.57,
                        "Taxa R$": 168.88,
                        "Taxa%": 6.00
                    },
                    {
                        "name": "SODEXO ALIMENTACAO",
                        "Venda Bruta": 251947.62,
                        "Taxa R$": 15117.24,
                        "Taxa%": 6.00
                    },
                    {
                        "name": "ALELO MULTIBENEFÍCIOS",
                        "Venda Bruta": 957.47,
                        "Taxa R$": 62.24,
                        "Taxa%": 6.50
                    },
                    {
                        "name": "TICKET ALIMENTACAO",
                        "Venda Bruta": 37257.34,
                        "Taxa R$": 1956.08,
                        "Taxa%": 5.25
                    },
                    {
                        "name": "TICKET REFEICAO",
                        "Venda Bruta": 3112.40,
                        "Taxa R$": 196.09,
                        "Taxa%": 6.30
                    },
                    {
                        "name": "VR ALIMENTACAO",
                        "Venda Bruta": 32413.00,
                        "Taxa R$": 2042.06,
                        "Taxa%": 6.30
                    },
                    {
                        "name": "VR REFEICAO",
                        "Venda Bruta": 1812.07,
                        "Taxa R$": 114.16,
                        "Taxa%": 6.30
                    },
                    {
                        "name": "TICKET FLEX",
                        "Venda Bruta": 1259.37,
                        "Taxa R$": 66.12,
                        "Taxa%": 5.25
                    },
                    {
                        "name": "VALESHOP DÉBITO",
                        "Venda Bruta": 10.38,
                        "Taxa R$": 0.62,
                        "Taxa%": 5.97
                    },
                    {
                        "name": "CABAL VOUCHER",
                        "Venda Bruta": 698.46,
                        "Taxa R$": 20.96,
                        "Taxa%": 3.00
                    }
                ]
            }
        ],
        "adquirente": [
            {
                "name": "Alelo",
                "Venda Bruta": 152172.32,
                "Soma_de_Valor_Taxa": "R$ 9.869,38",
                "Taxa%": 6.49
            },
            {
                "name": "Cabal",
                "Venda Bruta": 698.46,
                "Soma_de_Valor_Taxa": "R$ 20,96",
                "Taxa%": 3.00
            },
            {
                "name": "Cielo",
                "Venda Bruta": 309304.62,
                "Soma_de_Valor_Taxa": "R$ 11.449,97",
                "Taxa%": 3.70
            },
            {
                "name": "Getnet",
                "Venda Bruta": 1012888.90,
                "Soma_de_Valor_Taxa": "R$ 12.651,28",
                "Taxa%": 1.25
            },
            {
                "name": "Sodexo",
                "Venda Bruta": 255310.06,
                "Soma_de_Valor_Taxa": "R$ 15.319,00",
                "Taxa%": 6.00
            },
            {
                "name": "Ticket",
                "Venda Bruta": 41629.11,
                "Soma_de_Valor_Taxa": "R$ 2.218,29",
                "Taxa%": 5.33
            },
            {
                "name": "Tricard",
                "Venda Bruta": 16081.45,
                "Soma_de_Valor_Taxa": "R$ 514,58",
                "Taxa%": 3.20
            },
            {
                "name": "Valecard",
                "Venda Bruta": 13688.32,
                "Soma_de_Valor_Taxa": "R$ 889,79",
                "Taxa%": 6.50
            },
            {
                "name": "Valeshop",
                "Venda Bruta": 10.38,
                "Soma_de_Valor_Taxa": "R$ 0,62",
                "Taxa%": 5.97
            },
            {
                "name": "VR",
                "Venda Bruta": 34225.07,
                "Soma_de_Valor_Taxa": "R$ 2.156,22",
                "Taxa%": 6.30
            }
        ],
        "servicosAdicionaisPagos":[
            {
                "name": "Total",
                "Valor": 10328
            },
            {
                "name": "ANUIDADE (TICKET)",
                "Valor": 9364
            },
            {
                "name": "VALOR DA ANUIDADE (VR)",
                "Valor": 385
            },
            {
                "name": "(Em branco)",
                "Valor": 262
            },
            {
                "name": "HELP DESK (SODEXO)",
                "Valor": 180
            },                                                
            {
                "name": "TARIFA BANCÁRIA (TICKET & VR)",
                "Valor": 137
            }            
        ],
        "vendaPorTipoCartao":[
            {
                "id": 1,
                "title": "Débito",
                "value": 42.04 
            },
            {
                "id": 2,
                "title": "Crédito",
                "value": 30.87
            },
            {
                "id": 3,
                "title": "Voucher",
                "value": 27.09
            },                               
        ],        
    }
    
    return JsonResponse(analytics_data, safe=False)
